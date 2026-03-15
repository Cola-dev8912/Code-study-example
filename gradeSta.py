# -*- coding: utf-8 -*-
"""
统计工作表中「优、良、中、次、差」的个数及各自所占比例。
适用于路面损坏状况评定明细表中的等级列。
"""

import os
import sys

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 要统计的等级字符
GRADE_CHARS = ("优", "良", "中", "次", "差")
# 支持的 Excel 扩展名（openpyxl 可读取 .xlsx 与 .xlsm）
EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
# 要筛选的工作表指标（需同时满足：含指标名 且 含"百"）
INDICATORS = (
    "PQI", "PCI", "RQI", "RDI", "PBI", "SRI", "PSSI"
)


def count_grades_in_sheet(ws):
    """遍历工作表中所有单元格，统计等级字符出现次数。"""
    counts = {c: 0 for c in GRADE_CHARS}
    total = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val:
                continue
            # 只统计单字符或单元格整体为等级字的情况
            for char in val:
                if char in GRADE_CHARS:
                    counts[char] += 1
                    total += 1

    return counts, total


def _resolve_excel_path(path):
    """若指定路径不存在，尝试同名但另一扩展名（.xlsx / .xlsm）。"""
    if os.path.isfile(path):
        return path
    base, ext = os.path.splitext(path)
    for other in EXCEL_EXTENSIONS:
        if other != ext.lower():
            alt = base + other
            if os.path.isfile(alt):
                return alt
    return None


def _is_target_sheet(name):
    """
    判断工作表名是否需要统计。
    条件：名字中含有 INDICATORS 之一，且名字中含有"百"。
    例如："PCI百上" ✔，"RDI下" ✘，"PCI千下" ✘，"DCB百上二" ✘
    """
    name_upper = name.upper()
    has_indicator = any(ind in name_upper for ind in INDICATORS)
    has_bai = "百" in name
    return has_indicator and has_bai


def _extract_indicator(name):
    """从工作表名中提取指标名（如 "PQI百上" -> "PQI"）"""
    name_upper = name.upper()
    for ind in INDICATORS:
        if ind in name_upper:
            return ind
    return None


def main():
    # 默认当前目录下第一个 .xlsx/.xlsm 文件，或通过命令行参数指定
    if len(sys.argv) >= 2:
        path = sys.argv[1]
        resolved = _resolve_excel_path(path)
        if resolved and resolved != path:
            print(f"未找到 {path}，使用同名的 {resolved}\n")
            path = resolved
    else:
        candidates = [
            f for f in os.listdir(".")
            if any(f.lower().endswith(ext) for ext in EXCEL_EXTENSIONS)
            and not f.startswith("~$")
        ]
        if not candidates:
            print("当前目录下未找到 .xlsx / .xlsm 文件，请将 Excel 文件放入本目录或运行：")
            print("  python grade_stats.py <文件路径>")
            sys.exit(1)
        path = candidates[0]
        print(f"使用文件: {path}\n")

    if not os.path.isfile(path):
        print(f"文件不存在: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)

    # 筛选出需要统计的工作表
    target_sheets = [ws for ws in wb.worksheets if _is_target_sheet(ws.title)]

    if not target_sheets:
        print("未找到符合条件的工作表（名字含 PQI/PCI/RQI/RDI/PBI/SRI/PSSI 且含'百'）")
        wb.close()
        sys.exit(0)

    # 按指标分组统计
    indicator_data = {}  # {指标名: {"counts": {}, "total": 0, "sheets": []}}

    for ws in target_sheets:
        indicator = _extract_indicator(ws.title)
        if indicator is None:
            continue

        counts, total = count_grades_in_sheet(ws)

        if indicator not in indicator_data:
            indicator_data[indicator] = {
                "counts": {c: 0 for c in GRADE_CHARS},
                "total": 0,
                "sheets": []
            }

        for c in GRADE_CHARS:
            indicator_data[indicator]["counts"][c] += counts[c]
        indicator_data[indicator]["total"] += total
        indicator_data[indicator]["sheets"].append(ws.title)

    # 生成结果 Excel 文件
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "统计结果"

    # 定义样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    indicator_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 写入表头
    headers = ["指标", "工作表", "优", "良", "中", "次", "差", "合计", "优%","良%","中%","次%","差%"]
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 设置列宽
    col_widths = [10, 30, 8, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8]
    for col, width in enumerate(col_widths, 1):
        ws_out.column_dimensions[chr(64 + col)].width = width

    # 写入数据（按指标分组）
    row = 2
    for indicator in INDICATORS:
        if indicator not in indicator_data:
            continue

        data = indicator_data[indicator]
        counts = data["counts"]
        total = data["total"]
        sheets = data["sheets"]
        sheet_names = " + ".join(sheets)

        # 计算比例
        pct = {c: (counts[c] / total * 100) if total else 0 for c in GRADE_CHARS}

        # 写入该指标所在行
        ws_out.cell(row=row, column=1, value=indicator).fill = indicator_fill
        ws_out.cell(row=row, column=2, value=sheet_names)
        ws_out.cell(row=row, column=3, value=counts["优"])
        ws_out.cell(row=row, column=4, value=counts["良"])
        ws_out.cell(row=row, column=5, value=counts["中"])
        ws_out.cell(row=row, column=6, value=counts["次"])
        ws_out.cell(row=row, column=7, value=counts["差"])
        ws_out.cell(row=row, column=8, value=total)
        ws_out.cell(row=row, column=9, value=f'{pct["优"]:.1f}%')
        ws_out.cell(row=row, column=10, value=f'{pct["良"]:.1f}%')
        ws_out.cell(row=row, column=11, value=f'{pct["中"]:.1f}%')
        ws_out.cell(row=row, column=12, value=f'{pct["次"]:.1f}%')
        ws_out.cell(row=row, column=13, value=f'{pct["差"]:.1f}%')

        # 应用样式
        for col in range(1, 14):
            cell = ws_out.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align

        row += 1

    # 保存文件
    output_path = os.path.splitext(os.path.basename(path))[0] + "_统计结果.xlsx"
    wb_out.save(output_path)
    print(f"统计完成！结果已保存至: {output_path}")

    wb.close()


if __name__ == "__main__":
    main()
